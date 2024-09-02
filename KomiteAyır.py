import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import tkinter as tk
from tkinter import filedialog, ttk, messagebox

def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry_excel_path.delete(0, tk.END)
        entry_excel_path.insert(0, file_path)
        update_comboboxes(file_path)

def update_comboboxes(file_path):
    df = pd.read_excel(file_path)
    columns = list(df.columns)
    combo_komite_column['values'] = columns
    combo_sira_column['values'] = columns
    combo_range_start['values'] = list(range(1, 42))
    combo_range_end['values'] = list(range(1, 42))

def start_process():
    file_path = entry_excel_path.get()
    komite_column = combo_komite_column.get()
    sira_column = combo_sira_column.get()
    
    if not file_path or not komite_column or not sira_column:
        messagebox.showwarning("Uyarı", "Lütfen tüm seçimleri yapın.")
        return
    
    df = pd.read_excel(file_path)

    if var_range_select.get() == 1:
        start = int(combo_range_start.get())
        end = int(combo_range_end.get())
        komite_range = range(start, end + 1)
    else:
        komite_range = range(1, 42)

    progress_bar['maximum'] = len(komite_range)
    progress_bar['value'] = 0

    for idx, komite_numarasi in enumerate(komite_range, 1):
        komite_df = df[df[komite_column] == f"{komite_numarasi}. komite"].copy()
        
        if komite_df.empty:
            continue
        
        komite_df[sira_column] = range(1, len(komite_df) + 1)

        yeni_wb = load_workbook(file_path)
        yeni_ws = yeni_wb.active

        yeni_ws.delete_rows(2, yeni_ws.max_row)

        for r_idx, row in enumerate(dataframe_to_rows(komite_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = yeni_ws.cell(row=r_idx, column=c_idx, value=value)
                
                original_cell = yeni_ws.cell(row=r_idx, column=c_idx)
                cell.font = copy(original_cell.font)
                cell.border = copy(original_cell.border)
                cell.fill = copy(original_cell.fill)
                cell.number_format = copy(original_cell.number_format)
                cell.protection = copy(original_cell.protection)
                cell.alignment = copy(original_cell.alignment)
        
        max_row = yeni_ws.max_row
        for row in range(max_row, 1, -1):
            if all(yeni_ws.cell(row=row, column=c).value in [None, ""] for c in range(1, yeni_ws.max_column + 1)):
                yeni_ws.delete_rows(row, 1)
        
        dosya_adi = f"Komite_{komite_numarasi}.xlsx"
        dosya_yolu = os.path.join(os.path.expanduser("~"), "Desktop", "Komite_Dosyaları", dosya_adi)
        os.makedirs(os.path.dirname(dosya_yolu), exist_ok=True)
        yeni_wb.save(dosya_yolu)
        lbl_status.config(text=f"İşlem yapılıyor: {komite_numarasi}. komite")

        # İlerleme çubuğunu güncelle
        progress_bar['value'] = idx
        root.update_idletasks()

    lbl_status.config(text="Tüm komiteler için dosyalar oluşturuldu.")
    messagebox.showinfo("Bilgi", "Tüm komiteler için dosyalar oluşturuldu.")
    
    # İlerleme çubuğunu sıfırla
    progress_bar['value'] = 0

root = tk.Tk()
root.title("Komite Dosyası Oluşturucu")

frame = tk.Frame(root)
frame.pack(pady=20)

label_excel_path = tk.Label(frame, text="Excel Dosyası Seçin:")
label_excel_path.grid(row=0, column=0, padx=10, pady=5)
entry_excel_path = tk.Entry(frame, width=50)
entry_excel_path.grid(row=0, column=1, padx=10, pady=5)
button_browse = tk.Button(frame, text="Gözat", command=select_excel_file)
button_browse.grid(row=0, column=2, padx=10, pady=5)

label_komite_column = tk.Label(frame, text="Böl:")
label_komite_column.grid(row=1, column=0, padx=10, pady=5)
combo_komite_column = ttk.Combobox(frame, state="readonly")
combo_komite_column.grid(row=1, column=1, padx=10, pady=5)

label_sira_column = tk.Label(frame, text="Sıra Sütunu:")
label_sira_column.grid(row=2, column=0, padx=10, pady=5)
combo_sira_column = ttk.Combobox(frame, state="readonly")
combo_sira_column.grid(row=2, column=1, padx=10, pady=5)

var_range_select = tk.IntVar()
radio_all = tk.Radiobutton(frame, text="Tüm Komiteler (1-41)", variable=var_range_select, value=0)
radio_all.grid(row=3, column=0, padx=10, pady=5)
radio_range = tk.Radiobutton(frame, text="Komite Aralığı Seçin", variable=var_range_select, value=1)
radio_range.grid(row=3, column=1, padx=10, pady=5)

label_range_start = tk.Label(frame, text="Başlangıç Komitesi:")
label_range_start.grid(row=4, column=0, padx=10, pady=5)
combo_range_start = ttk.Combobox(frame, state="readonly")
combo_range_start.grid(row=4, column=1, padx=10, pady=5)

label_range_end = tk.Label(frame, text="Bitiş Komitesi:")
label_range_end.grid(row=5, column=0, padx=10, pady=5)
combo_range_end = ttk.Combobox(frame, state="readonly")
combo_range_end.grid(row=5, column=1, padx=10, pady=5)

button_start = tk.Button(root, text="Başlat", command=start_process)
button_start.pack(pady=10)

progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.pack(pady=10)

lbl_status = tk.Label(root, text="")
lbl_status.pack(pady=10)

root.mainloop()
